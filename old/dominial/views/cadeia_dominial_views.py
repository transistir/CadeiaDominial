from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from ..models import Imovel, TIs, Documento, Lancamento, Cartorios, DocumentoTipo
from ..services import HierarquiaService
from ..services.hierarquia_arvore_service import HierarquiaArvoreService
from ..services.cache_service import CacheService
from ..services.cadeia_dominial_tabela_service import CadeiaDominialTabelaService
from datetime import date
import json
from weasyprint import HTML
from django.template.loader import render_to_string
from django.conf import settings
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@login_required
def cadeia_dominial(request, tis_id, imovel_id):
    # Otimização: usar select_related para reduzir queries
    tis = get_object_or_404(TIs, id=tis_id)
    imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
    
    # Obter documentos seguindo a hierarquia correta (tronco principal)
    tronco_principal = HierarquiaService.obter_tronco_principal(imovel)
    documentos = list(tronco_principal)  # Usar a ordem do tronco principal
    
    tem_documentos = len(documentos) > 0
    tem_apenas_matricula = len(documentos) == 1 and documentos[0].tipo.tipo == 'matricula' if tem_documentos else False
    
    # Otimização: usar exists() em vez de count() para verificar lançamentos
    tem_lancamentos = Lancamento.objects.filter(documento__imovel=imovel).exists() if tem_documentos else False
    mostrar_lancamentos = request.GET.get('lancamentos') == 'true'

    # Refatoração: delegar identificação de troncos para o service
    troncos_secundarios = []
    if tem_documentos:
        troncos_secundarios = HierarquiaService.obter_troncos_secundarios(imovel)

    context = {
        'tis': tis,
        'imovel': imovel,
        'documentos': documentos,
        'tem_apenas_matricula': tem_apenas_matricula,
        'tem_lancamentos': tem_lancamentos,
        'mostrar_lancamentos': mostrar_lancamentos,
        'tronco_principal': tronco_principal,
        'troncos_secundarios': troncos_secundarios,
    }

    if mostrar_lancamentos:
        return render(request, 'dominial/cadeia_dominial.html', context)
    else:
        return render(request, 'dominial/cadeia_dominial_arvore.html', context)

@login_required
def cadeia_dominial_arvore(request, tis_id, imovel_id):
    """Retorna os dados da cadeia dominial em formato de árvore para o diagrama"""
    try:
        tis = get_object_or_404(TIs, id=tis_id)
        imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
        # Delegar a construção da árvore para um service/utilitário
        arvore = HierarquiaArvoreService.construir_arvore_cadeia_dominial(imovel, criar_documentos_automaticos=True)
        
        # Adicionar headers para evitar cache
        response = JsonResponse(arvore, safe=False)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def tronco_principal(request, tis_id, imovel_id):
    """Exibe o tronco principal da cadeia dominial em formato de tabela"""
    tis = get_object_or_404(TIs, id=tis_id)
    imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
    
    # Obter escolhas de origem da URL (se houver)
    escolhas_origem = {}
    escolhas_param = request.GET.get('escolhas')
    if escolhas_param:
        try:
            escolhas_origem = json.loads(escolhas_param)
        except json.JSONDecodeError:
            escolhas_origem = {}
    
    # Obter cadeia em formato de tabela
    service = CadeiaDominialTabelaService()
    
    # Se há escolhas de origem, usar o método completo
    if escolhas_origem:
        result = service.get_cadeia_dominial_tabela(tis_id, imovel_id, request.session, escolhas_origem)
        cadeia = result.get('cadeia', [])
    else:
        # Se não há escolhas, usar o método simples (tronco principal)
        cadeia = service.obter_cadeia_tabela(imovel, escolhas_origem)
    
    # Verificar se há lançamentos
    tem_lancamentos = False
    if cadeia:
        tem_lancamentos = any(len(item['lancamentos']) > 0 for item in cadeia)
    
    context = {
        'tis': tis,
        'imovel': imovel,
        'cadeia': cadeia,
        'tem_lancamentos': tem_lancamentos,
        'tipo_visualizacao': 'tabela',
    }
    
    return render(request, 'dominial/cadeia_dominial_tabela.html', context)

@login_required
def cadeia_dominial_dados(request, tis_id, imovel_id):
    """Retorna os dados da cadeia dominial em formato JSON para o diagrama de árvore"""
    tis = get_object_or_404(TIs, id=tis_id)
    imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
    
    # Obter documentos seguindo a hierarquia correta (tronco principal)
    tronco_principal = HierarquiaService.obter_tronco_principal(imovel)
    documentos = list(tronco_principal)  # Usar a ordem do tronco principal
    
    # Estrutura para o diagrama de árvore
    tree_data = {
        'name': f'Imóvel: {imovel.matricula}',
        'children': []
    }
    
    for documento in documentos:
        doc_node = {
            'name': f'{documento.tipo.get_tipo_display()}: {documento.numero}',
            'data': {
                'tipo': 'documento',
                'id': documento.id,
                'numero': documento.numero,
                'data': documento.data.strftime('%d/%m/%Y'),
                'cartorio': documento.cartorio.nome,
                'livro': documento.livro,
                'folha': documento.folha,
                'origem': documento.origem or '',
                'tipo_documento': documento.tipo.tipo,  # Adicionar tipo do documento
                'classificacao_fim_cadeia': documento.classificacao_fim_cadeia,
                'sigla_patrimonio_publico': documento.sigla_patrimonio_publico,  # Adicionar classificação de fim de cadeia
                'sigla_patrimonio_publico': documento.sigla_patrimonio_publico  # Adicionar sigla do patrimônio público
            },
            'children': []
        }
        
        # Adicionar lançamentos como filhos do documento
        # Otimização: usar a lista já pré-carregada
        lancamentos = list(documento.lancamentos.all())
        lancamentos.sort(key=lambda x: x.id)  # Ordenar por ID
        
        for lancamento in lancamentos:
            lanc_node = {
                'name': f'{lancamento.tipo.get_tipo_display()}: {lancamento.data.strftime("%d/%m/%Y")}',
                'data': {
                    'tipo': 'lancamento',
                    'id': lancamento.id,
                    'tipo_lancamento': lancamento.tipo.tipo,
                    'data': lancamento.data.strftime('%d/%m/%Y'),
                    'eh_inicio_matricula': lancamento.eh_inicio_matricula,
                    'forma': lancamento.forma or '',
                    'descricao': lancamento.descricao or '',
                    'titulo': lancamento.titulo or '',
                    'observacoes': lancamento.observacoes or ''
                }
            }
            doc_node['children'].append(lanc_node)
        
        tree_data['children'].append(doc_node)
    
    return JsonResponse(tree_data, safe=False)

@login_required
def cadeia_dominial_tabela(request, tis_id, imovel_id):
    """
    View para visualização de tabela da cadeia dominial
    """
    # NOVO: processar escolha de origem
    origem_escolhida = request.GET.get('origem')
    documento_id = request.GET.get('documento_id')
    if origem_escolhida and documento_id:
        request.session[f'origem_documento_{documento_id}'] = origem_escolhida

    service = CadeiaDominialTabelaService()
    context = service.get_cadeia_dominial_tabela(tis_id, imovel_id, request.session)

    # Adicionar estatísticas
    if context['cadeia']:
        context['estatisticas'] = service.get_estatisticas_cadeia(context['cadeia'])

    return render(request, 'dominial/cadeia_dominial_tabela.html', context)

@login_required
def cadeia_dominial_d3(request, tis_id, imovel_id):
    """Nova visualização D3.js da árvore da cadeia dominial"""
    tis = get_object_or_404(TIs, id=tis_id)
    imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
    documentos = Documento.objects.filter(imovel=imovel)\
        .select_related('cartorio', 'tipo')\
        .prefetch_related('lancamentos', 'lancamentos__tipo')\
        .order_by('data')
    tem_documentos = documentos.exists()
    tem_apenas_matricula = documentos.count() == 1 and documentos.first().tipo.tipo == 'matricula' if tem_documentos else False
    tem_lancamentos = Lancamento.objects.filter(documento__imovel=imovel).exists() if tem_documentos else False
    context = {
        'tis': tis,
        'imovel': imovel,
        'documentos': documentos,
        'tem_apenas_matricula': tem_apenas_matricula,
        'tem_lancamentos': tem_lancamentos,
    }
    return render(request, 'dominial/cadeia_dominial_d3.html', context) 

@login_required
def documento_detalhado(request, tis_id, imovel_id, documento_id):
    """
    View para visualização detalhada de um documento específico
    Baseada na cadeia dominial tabela, mas mostra apenas um documento
    Suporta documentos importados de outras cadeias dominiais
    """
    tis = get_object_or_404(TIs, id=tis_id)
    imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
    
    # Buscar o documento - pode estar em outro imóvel se for importado
    documento = None
    is_importado = False
    cadeias_dominiais = []
    
    # Primeiro, tentar encontrar no imóvel atual
    try:
        documento = Documento.objects.get(id=documento_id, imovel=imovel)
    except Documento.DoesNotExist:
        # Se não encontrou no imóvel atual, pode ser um documento importado
        try:
            documento = Documento.objects.get(id=documento_id)
            is_importado = True
            
            # Buscar informações de importação
            from ..models import DocumentoImportado
            importacoes = DocumentoImportado.objects.filter(documento=documento)
            for importacao in importacoes:
                cadeias_dominiais.append({
                    'imovel_id': importacao.imovel_origem.id,
                    'imovel_matricula': importacao.imovel_origem.matricula,
                    'imovel_nome': importacao.imovel_origem.nome,
                    'data_importacao': importacao.data_importacao.strftime('%d/%m/%Y'),
                    'importado_por': importacao.importado_por.username if importacao.importado_por else 'Sistema'
                })
        except Documento.DoesNotExist:
            # Documento não existe
            from django.http import Http404
            raise Http404("Documento não encontrado")
    
    # Carregar lançamentos do documento
    lancamentos = documento.lancamentos.select_related('tipo').prefetch_related(
        'pessoas__pessoa'
    ).order_by('id')
    
    # Aplicar ordenação personalizada por número simples (decrescente)
    from ..services.lancamento_consulta_service import LancamentoConsultaService
    lancamentos_list = list(lancamentos)
    lancamentos_list.sort(key=lambda x: (
        -LancamentoConsultaService._extrair_numero_simples(x.numero_lancamento),
        x.id
    ))
    lancamentos = lancamentos_list
    
    context = {
        'tis': tis,
        'imovel': imovel,
        'documento': documento,
        'lancamentos': lancamentos,
        'tipo_visualizacao': 'documento_unico',
        'tem_lancamentos': len(lancamentos) > 0,
        'is_importado': is_importado,
        'cadeias_dominiais': cadeias_dominiais,
        'total_cadeias': len(cadeias_dominiais)
    }
    
    return render(request, 'dominial/documento_detalhado.html', context) 

@login_required
def exportar_cadeia_dominial_pdf(request, tis_id, imovel_id):
    """
    Exporta a cadeia dominial em formato PDF
    """
    try:
        tis = get_object_or_404(TIs, id=tis_id)
        imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
        
        # Obter dados da cadeia dominial
        service = CadeiaDominialTabelaService()
        context = service.get_cadeia_dominial_tabela(tis_id, imovel_id, request.session)
        
        # Adicionar estatísticas
        if context['cadeia']:
            context['estatisticas'] = service.get_estatisticas_cadeia(context['cadeia'])
        
        # Renderizar template HTML para PDF
        html_string = render_to_string('dominial/cadeia_dominial_pdf.html', context)
        
        # Configurar CSS para PDF
        css_path = os.path.join(settings.STATIC_ROOT, 'dominial', 'css', 'cadeia_dominial_pdf.css')
        if not os.path.exists(css_path):
            css_path = os.path.join(settings.STATICFILES_DIRS[0], 'dominial', 'css', 'cadeia_dominial_pdf.css')
        
        # Gerar PDF
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(
            stylesheets=[css_path] if os.path.exists(css_path) else None
        )
        
        # Configurar resposta HTTP
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"cadeia_dominial_{imovel.matricula}_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        # Em caso de erro, retornar uma página de erro simples
        error_html = f"""
        <html>
        <head><title>Erro na Geração do PDF</title></head>
        <body>
            <h1>Erro na Geração do PDF</h1>
            <p>Ocorreu um erro ao gerar o PDF da cadeia dominial.</p>
            <p>Erro: {str(e)}</p>
            <p><a href="javascript:history.back()">Voltar</a></p>
        </body>
        </html>
        """
        return HttpResponse(error_html, content_type='text/html')

@login_required
def exportar_cadeia_completa_pdf(request, tis_id, imovel_id):
    try:
        tis = get_object_or_404(TIs, id=tis_id)
        imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
        
        # Verificar se há sequência personalizada
        sequencia_ids = request.GET.get('sequencia')
        
        from ..services.cadeia_completa_service import CadeiaCompletaService
        service = CadeiaCompletaService()
        
        if sequencia_ids:
            # Usar sequência personalizada
            context = service.get_cadeia_completa_com_sequencia_personalizada(tis_id, imovel_id, sequencia_ids)
        else:
            # Usar sequência padrão
            context = service.get_cadeia_completa(tis_id, imovel_id)
        
        html_string = render_to_string('dominial/cadeia_completa_pdf.html', context)
        css_path = os.path.join(settings.STATIC_ROOT, 'dominial', 'css', 'cadeia_completa_pdf.css')
        if not os.path.exists(css_path):
            css_path = os.path.join(settings.STATICFILES_DIRS[0], 'dominial', 'css', 'cadeia_completa_pdf.css')
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(
            stylesheets=[css_path] if os.path.exists(css_path) else None
        )
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"cadeia_completa_{imovel.matricula}_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        error_html = f"""
        <html>
        <head><title>Erro na Geração do PDF</title></head>
        <body style="font-family: Arial, sans-serif; padding: 20px; background-color: #f8f9fa;">
            <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                <h1 style="color: #dc3545; margin-bottom: 20px;">❌ Erro na Geração do PDF</h1>
                <p style="color: #6c757d; margin-bottom: 15px;">Ocorreu um erro ao gerar o PDF da cadeia dominial completa.</p>
                <div style="background-color: #f8f9fa; padding: 15px; border-radius: 4px; border-left: 4px solid #dc3545;">
                    <strong>Erro:</strong> {str(e)}
                </div>
                <div style="margin-top: 20px;">
                    <a href="javascript:history.back()" style="background-color: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px;">← Voltar</a>
                </div>
            </div>
        </body>
        </html>
        """
        return HttpResponse(error_html, content_type='text/html')

@login_required
def exportar_cadeia_dominial_excel(request, tis_id, imovel_id):
    """
    Exporta a cadeia dominial geral em formato Excel (mesma estrutura da página ver-cadeia-dominial)
    """
    try:
        tis = get_object_or_404(TIs, id=tis_id)
        imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
        
        # Usar o CadeiaCompletaService (mesmo do PDF) para incluir TODOS os documentos
        from ..services.cadeia_completa_service import CadeiaCompletaService
        service = CadeiaCompletaService()
        context = service.get_cadeia_completa(tis_id, imovel_id)
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Cadeia Dominial Geral"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalho principal
        ws.merge_cells('A1:P1')
        ws['A1'] = f"CADEIA DOMINIAL GERAL - {imovel.nome}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        
        # Informações do imóvel
        ws['A3'] = "TIS:"
        ws['B3'] = tis.nome
        ws['A4'] = "Matrícula:"
        ws['B4'] = imovel.matricula
        ws['A5'] = "Nome:"
        ws['B5'] = imovel.nome
        ws['A6'] = "Proprietário:"
        ws['B6'] = imovel.proprietario.nome if imovel.proprietario else ""
        ws['A7'] = "Cartório:"
        ws['B7'] = imovel.cartorio.nome if imovel.cartorio else ""
        ws['A8'] = "Data de Exportação:"
        ws['B8'] = date.today().strftime('%d/%m/%Y')
        
        # Inicializar linha atual
        row = 10
        
        # Processar a cadeia completa (mesma estrutura do PDF)
        cadeia_completa = context['cadeia_completa']
        
        for tronco in cadeia_completa:
            # Processar documentos do tronco
            for item in tronco['documentos']:
                documento = item['documento']
                lancamentos = item['lancamentos']
                is_importado = item.get('is_importado', False)
            
                # Título do documento
                row += 1
                prefixo_importado = "📥 " if is_importado else ""
                ws.merge_cells(f'A{row}:P{row}')
                ws.cell(row=row, column=1, value=f"{prefixo_importado}{documento.tipo.get_tipo_display()}: {documento.numero}").font = Font(bold=True, size=12)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
                ws.cell(row=row, column=1).alignment = center_alignment
                row += 1
            
                # Cabeçalho da tabela de lançamentos (igual ao template)
                # Primeira linha de cabeçalho (agrupamentos)
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row=row, column=1, value="MATRÍCULA").font = header_font
                ws.cell(row=row, column=1).fill = header_fill
                ws.cell(row=row, column=1).alignment = center_alignment
                
                ws.merge_cells(f'F{row}:G{row}')
                ws.cell(row=row, column=6, value="").fill = header_fill
                
                ws.merge_cells(f'H{row}:M{row}')
                ws.cell(row=row, column=8, value="TRANSMISSÃO").font = header_font
                ws.cell(row=row, column=8).fill = header_fill
                ws.cell(row=row, column=8).alignment = center_alignment
                
                ws.cell(row=row, column=14, value="Área (ha)").font = header_font
                ws.cell(row=row, column=14).fill = header_fill
                ws.cell(row=row, column=14).alignment = center_alignment
                
                ws.cell(row=row, column=15, value="Origem").font = header_font
                ws.cell(row=row, column=15).fill = header_fill
                ws.cell(row=row, column=15).alignment = center_alignment
                
                ws.cell(row=row, column=16, value="Observações").font = header_font
                ws.cell(row=row, column=16).fill = header_fill
                ws.cell(row=row, column=16).alignment = center_alignment
                row += 1
            
                # Segunda linha de cabeçalho (colunas específicas)
                headers_detalhados = [
                    'Nº', 'L', 'Fls.', 'Cartório', 'Data',  # Matrícula
                    'Transmitente', 'Adquirente',  # Pessoas
                    'Forma', 'Título', 'Cartório', 'L', 'Fls.', 'Data',  # Transmissão
                    'Área (ha)', 'Origem', 'Observações'
                ]
                
                for col, header in enumerate(headers_detalhados, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                row += 1
            
                # Adicionar lançamentos
                for lancamento in lancamentos:
                    # Nº (usando o filtro numero_documento_criado)
                    from ..templatetags.dominial_extras import numero_documento_criado
                    numero_formatado = numero_documento_criado(lancamento)
                    ws.cell(row=row, column=1, value=numero_formatado).border = border
                    
                    # L, Fls., Cartório, Data (do documento)
                    ws.cell(row=row, column=2, value=documento.livro or "-").border = border
                    ws.cell(row=row, column=3, value=documento.folha or "-").border = border
                    ws.cell(row=row, column=4, value=documento.cartorio.nome if documento.cartorio else "-").border = border
                    ws.cell(row=row, column=5, value=lancamento.data.strftime('%d/%m/%Y') if lancamento.data else "-").border = border
                    
                    # Transmitente
                    transmitentes = [p.pessoa.nome for p in lancamento.transmitentes.all()]
                    ws.cell(row=row, column=6, value=", ".join(transmitentes) if transmitentes else "-").border = border
                    
                    # Adquirente
                    adquirentes = [p.pessoa.nome for p in lancamento.adquirentes.all()]
                    ws.cell(row=row, column=7, value=", ".join(adquirentes) if adquirentes else "-").border = border
                    
                    # Transmissão
                    if lancamento.tipo.tipo == 'averbacao':
                        # Para averbação, mesclar colunas e mostrar descrição
                        ws.merge_cells(f'H{row}:M{row}')
                        ws.cell(row=row, column=8, value=lancamento.descricao or "-").border = border
                    else:
                        # Para outros tipos, mostrar campos específicos
                        ws.cell(row=row, column=8, value=lancamento.forma or "-").border = border
                        ws.cell(row=row, column=9, value=lancamento.titulo or "-").border = border
                        ws.cell(row=row, column=10, value=lancamento.cartorio_transmissao_compat.nome if lancamento.cartorio_transmissao_compat else "-").border = border
                        ws.cell(row=row, column=11, value=lancamento.livro_transacao or "-").border = border
                        ws.cell(row=row, column=12, value=lancamento.folha_transacao or "-").border = border
                        ws.cell(row=row, column=13, value=lancamento.data_transacao.strftime('%d/%m/%Y') if lancamento.data_transacao else "-").border = border
                    
                    # Área, Origem, Observações
                    ws.cell(row=row, column=14, value="").border = border  # Área (não implementado)
                    ws.cell(row=row, column=15, value=lancamento.origem or "-").border = border
                    ws.cell(row=row, column=16, value=lancamento.observacoes or "-").border = border
                    
                    row += 1
                
                # Adicionar linha em branco entre documentos
                row += 1
        
        # Adicionar estatísticas (se disponíveis)
        if 'estatisticas' in context:
            row += 1
            estatisticas = context['estatisticas']
            
            # Título das estatísticas
            ws.merge_cells(f'A{row}:P{row}')
            ws.cell(row=row, column=1, value="📊 ESTATÍSTICAS").font = Font(bold=True, size=14, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            ws.cell(row=row, column=1).alignment = center_alignment
            row += 1
            
            # Estatísticas
            if 'total_documentos' in estatisticas:
                ws.cell(row=row, column=1, value="Total de Documentos:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=estatisticas['total_documentos']).border = border
                row += 1
            
            if 'total_lancamentos' in estatisticas:
                ws.cell(row=row, column=1, value="Total de Lançamentos:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=estatisticas['total_lancamentos']).border = border
                row += 1
            
            if 'documentos_compartilhados' in estatisticas:
                ws.cell(row=row, column=1, value="Documentos Compartilhados:").font = Font(bold=True)
                ws.cell(row=row, column=2, value=estatisticas['documentos_compartilhados']).border = border
                row += 1
        
        # Ajustar largura das colunas (16 colunas)
        column_widths = [12, 8, 8, 20, 12, 20, 20, 15, 15, 20, 8, 8, 12, 12, 20, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Configurar resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"cadeia_dominial_geral_{imovel.matricula}_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Salvar workbook
        wb.save(response)
        return response
        
    except Exception as e:
        # Em caso de erro, retornar uma resposta de erro
        error_response = HttpResponse(
            f"Erro ao gerar Excel: {str(e)}",
            content_type='text/plain'
        )
        error_response.status_code = 500
        return error_response

@login_required
def obter_arvore_cadeia_dominial(request, tis_id, imovel_id):
    """Retorna os dados da árvore da cadeia dominial para o modal de seleção de sequência"""
    try:
        tis = get_object_or_404(TIs, id=tis_id)
        imovel = get_object_or_404(Imovel, id=imovel_id, terra_indigena_id=tis)
        
        # 1. Primeiro obter o tronco principal na sequência correta
        tronco_principal = HierarquiaService.obter_tronco_principal(imovel)
        
        # 2. Obter todos os documentos da árvore
        arvore = HierarquiaService.construir_arvore_cadeia_dominial(imovel)
        
        # 3. Organizar documentos: tronco principal primeiro, depois o resto
        documentos_organizados = []
        documentos_processados_ids = set()
        
        # Primeiro, adicionar tronco principal
        for documento in tronco_principal:
            documento_processado = {
                'id': documento.id,
                'numero': documento.numero,
                'tipo': documento.tipo.tipo,
                'tipo_display': documento.tipo.get_tipo_display(),
                'data': documento.data.strftime('%d/%m/%Y'),
                'cartorio': documento.cartorio.nome,
                'livro': documento.livro or '',
                'folha': documento.folha or '',
                'origem': '',  # Tronco principal não tem origem específica
                'nivel': 0,  # Considerar tronco principal como nível 0
                'is_importado': documento.imovel != imovel,
                'is_compartilhado': False,
                'lancamentos_count': documento.lancamentos.count(),
                'detalhes': f"{documento.data.strftime('%d/%m/%Y')} - {documento.cartorio.nome}",
                'classificacao_fim_cadeia': documento.classificacao_fim_cadeia,
                'sigla_patrimonio_publico': documento.sigla_patrimonio_publico
            }
            documentos_organizados.append(documento_processado)
            documentos_processados_ids.add(documento.id)
        
        # Depois, organizar outros documentos seguindo lógica hierárquica
        outros_documentos = []
        for doc in arvore['documentos']:
            if doc['id'] not in documentos_processados_ids:
                documento_processado = {
                    'id': doc['id'],
                    'numero': doc['numero'],
                    'tipo': doc['tipo'],
                    'tipo_display': doc['tipo_display'],
                    'data': doc['data'],
                    'cartorio': doc['cartorio'],
                    'livro': doc['livro'],
                    'folha': doc['folha'],
                    'origem': doc['origem'],
                    'nivel': doc['nivel'],
                    'is_importado': doc.get('is_importado', False),
                    'is_compartilhado': doc.get('is_compartilhado', False),
                    'lancamentos_count': len(doc.get('lancamentos', [])),
                    'detalhes': f"{doc['data']} - {doc['cartorio']}",
                    'classificacao_fim_cadeia': doc.get('classificacao_fim_cadeia', None)
                }
                outros_documentos.append(documento_processado)
        
        # Organizar outros documentos seguindo lógica hierárquica
        outros_organizados = organizar_documentos_hierarquicamente(outros_documentos, arvore)
        
        # Adicionar outros documentos após o tronco principal
        documentos_organizados.extend(outros_organizados)
        
        response_data = {
            'success': True,
            'documentos': documentos_organizados,
            'total_documentos': len(documentos_organizados),
            'tronco_principal_count': len(tronco_principal),
            'imovel': {
                'id': imovel.id,
                'nome': imovel.nome,
                'matricula': imovel.matricula
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def organizar_documentos_hierarquicamente(documentos, arvore):
    """
    Organiza documentos seguindo lógica hierárquica:
    1. Maior número do menor nível
    2. Expandir origens do documento (maior número primeiro)
    3. Repetir até incluir todos os documentos
    """
    if not documentos:
        return []
    
    # Criar mapa de documentos por ID para facilitar busca
    docs_por_id = {doc['id']: doc for doc in documentos}
    
    # Criar mapa de conexões pai-filho
    conexoes = {}
    for con in arvore.get('conexoes', []):
        if con['from'] in docs_por_id and con['to'] in docs_por_id:
            if con['to'] not in conexoes:
                conexoes[con['to']] = []
            conexoes[con['to']].append(con['from'])
    
    # Organizar documentos por nível
    docs_por_nivel = {}
    for doc in documentos:
        nivel = doc['nivel']
        if nivel not in docs_por_nivel:
            docs_por_nivel[nivel] = []
        docs_por_nivel[nivel].append(doc)
    
    # Ordenar documentos em cada nível por número (maior primeiro)
    for nivel in docs_por_nivel:
        docs_por_nivel[nivel].sort(key=lambda x: 
            -int(x['numero'].replace('M', '').replace('T', '')) if x['numero'].replace('M', '').replace('T', '').isdigit() else 0
        )
    
    # Algoritmo de organização hierárquica
    documentos_organizados = []
    visitados = set()
    
    # Processar por níveis, do menor para o maior
    niveis_ordenados = sorted(docs_por_nivel.keys())
    
    for nivel in niveis_ordenados:
        docs_nivel = docs_por_nivel[nivel]
        
        for doc in docs_nivel:
            if doc['id'] not in visitados:
                # Adicionar documento atual
                documentos_organizados.append(doc)
                visitados.add(doc['id'])
                
                # Expandir origens deste documento
                expandir_origens_hierarquicamente(doc, conexoes, docs_por_id, documentos_organizados, visitados)
    
    return documentos_organizados

def expandir_origens_hierarquicamente(documento, conexoes, docs_por_id, documentos_organizados, visitados):
    """
    Expande origens de um documento seguindo lógica hierárquica
    """
    if documento['id'] not in conexoes:
        return
    
    # Obter origens do documento
    origens_ids = conexoes[documento['id']]
    origens = [docs_por_id[origem_id] for origem_id in origens_ids if origem_id in docs_por_id]
    
    # Ordenar origens por número (maior primeiro)
    origens.sort(key=lambda x: 
        -int(x['numero'].replace('M', '').replace('T', '')) if x['numero'].replace('M', '').replace('T', '').isdigit() else 0
    )
    
    # Adicionar origens na ordem
    for origem in origens:
        if origem['id'] not in visitados:
            documentos_organizados.append(origem)
            visitados.add(origem['id'])
            
            # Recursivamente expandir origens desta origem
            expandir_origens_hierarquicamente(origem, conexoes, docs_por_id, documentos_organizados, visitados) 