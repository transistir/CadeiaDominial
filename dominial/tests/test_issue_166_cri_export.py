"""
Issue #166 — usar a sigla "CRI" no lugar de "Cartório de Registro de Imóveis"
nas exportações (Excel e PDF completo).

Escopo: sigla nos valores + rótulos, APENAS nas exportações. A UI de cadastro
continua exibindo o nome por extenso (decisão da issue #50). Models,
migrations e dados não são alterados.
"""

import unicodedata
from io import BytesIO
from types import SimpleNamespace

from django.core.cache import cache
from django.template import Context, Template
from django.template.loader import render_to_string
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.utils import timezone
from openpyxl import load_workbook

from dominial.models import (
    Cartorios,
    Documento,
    DocumentoTipo,
    Imovel,
    Lancamento,
    LancamentoTipo,
    Pessoas,
    TIs,
)
from dominial.services.cadeia_completa_service import CadeiaCompletaService
from dominial.utils.formatacao_utils import abreviar_cartorio
from dominial.views import cadeia_dominial_views


class AbreviarCartorioTest(SimpleTestCase):
    """Testes unitários da função utilitária `abreviar_cartorio`."""

    def test_prefixo_simples(self):
        self.assertEqual(
            abreviar_cartorio("Cartório de Registro de Imóveis de Prado"),
            "CRI de Prado",
        )

    def test_prefixo_com_comarca(self):
        self.assertEqual(
            abreviar_cartorio(
                "Cartório de Registro de Imóveis da Comarca de Teodoro Sampaio"
            ),
            "CRI da Comarca de Teodoro Sampaio",
        )

    def test_nome_sem_prefixo_fica_inalterado(self):
        self.assertEqual(
            abreviar_cartorio("2º Ofício de Notas de Belém"),
            "2º Ofício de Notas de Belém",
        )

    def test_prefixo_em_caixa_alta_preserva_caixa_do_resto(self):
        self.assertEqual(
            abreviar_cartorio("CARTÓRIO DE REGISTRO DE IMÓVEIS DE PRADO"),
            "CRI DE PRADO",
        )

    def test_prefixo_sem_resto(self):
        self.assertEqual(
            abreviar_cartorio("Cartório de Registro de Imóveis"),
            "CRI",
        )

    def test_none_fica_inalterado(self):
        self.assertIsNone(abreviar_cartorio(None))

    def test_string_vazia_fica_inalterada(self):
        self.assertEqual(abreviar_cartorio(""), "")

    def test_variante_do_registro_nao_casa(self):
        self.assertEqual(
            abreviar_cartorio("Cartório do Registro de Imóveis de X"),
            "Cartório do Registro de Imóveis de X",
        )

    def test_texto_decomposto_nfd_e_abreviado(self):
        """Unicode decomposto (NFD) tem de abreviar igual ao NFC (issue #166 review)."""
        # "Cartório" e "Imóveis" com o acento agudo como caractere combinante
        # isolado (U+0301) — visualmente idêntico ao NFC, byte a byte diferente.
        nfd = "Cartório de Registro de Imóveis de Prado"
        self.assertEqual(nfd, unicodedata.normalize("NFD", nfd))
        self.assertEqual(abreviar_cartorio(nfd), "CRI de Prado")

    def test_nfc_e_nfd_produzem_mesmo_resultado(self):
        nfc = unicodedata.normalize(
            "NFC", "Cartório de Registro de Imóveis da Comarca de Teodoro Sampaio"
        )
        nfd = unicodedata.normalize("NFD", nfc)
        self.assertNotEqual(nfc, nfd)  # de fato são representações distintas
        self.assertEqual(
            abreviar_cartorio(nfc),
            abreviar_cartorio(nfd),
        )
        self.assertEqual(
            abreviar_cartorio(nfd), "CRI da Comarca de Teodoro Sampaio"
        )

    def test_sem_fronteira_de_palavra_nao_abrevia(self):
        """
        Exige fim de string ou caractere não-alfabético logo após "Imóveis"
        (issue #166 review). Comportamento escolhido: o nome volta inalterado.
        """
        original = "Cartório de Registro de ImóveisXYZ"
        self.assertEqual(abreviar_cartorio(original), original)

    def test_fronteira_com_pontuacao_ainda_abrevia(self):
        self.assertEqual(
            abreviar_cartorio("Cartório de Registro de Imóveis - Teodoro Sampaio"),
            "CRI - Teodoro Sampaio",
        )


class AbreviarCartorioTemplatetagTest(SimpleTestCase):
    """O filtro `abreviar_cartorio` registrado em `dominial_extras`."""

    def _render(self, valor):
        template = Template(
            "{% load dominial_extras %}{{ valor|abreviar_cartorio }}"
        )
        return template.render(Context({"valor": valor}))

    def test_filtro_abrevia_prefixo(self):
        self.assertEqual(
            self._render("Cartório de Registro de Imóveis de Prado"),
            "CRI de Prado",
        )

    def test_filtro_mantem_nome_sem_prefixo(self):
        self.assertEqual(
            self._render("2º Ofício de Notas de Belém"),
            "2º Ofício de Notas de Belém",
        )


class ExportacaoUsaSiglaCRITest(TestCase):
    """
    Excel e PDF completo devem trazer a sigla CRI nos valores tabulares e nos
    cabeçalhos das colunas de cartório, preservando o nome por extenso apenas
    no bloco de informações básicas do imóvel (texto narrativo).
    """

    NOME_CRI_IMOVEL = "Cartório de Registro de Imóveis de Prado"
    NOME_CRI_TRANSMISSAO = "Cartório de Registro de Imóveis de Teodoro Sampaio"

    def setUp(self):
        cache.clear()
        self.factory = RequestFactory()

        self.tis = TIs.objects.create(
            nome="TI Teste 166", codigo="TI166", etnia="Teste"
        )
        self.cartorio = Cartorios.objects.create(
            nome=self.NOME_CRI_IMOVEL,
            cns="166166",
            cidade="Prado",
            estado="BA",
        )
        self.cartorio_transmissao = Cartorios.objects.create(
            nome=self.NOME_CRI_TRANSMISSAO,
            cns="166167",
            cidade="Teodoro Sampaio",
            estado="SP",
        )
        self.proprietario = Pessoas.objects.create(
            nome="Proprietário 166", cpf="55566677788"
        )
        self.imovel = Imovel.objects.create(
            terra_indigena_id=self.tis,
            nome="Imóvel Teste 166",
            proprietario=self.proprietario,
            matricula="M700",
            cartorio=self.cartorio,
        )

        self.tipo_matricula = DocumentoTipo.objects.create(tipo="matricula")
        self.tipo_inicio = LancamentoTipo.objects.create(tipo="inicio_matricula")

        self.documento = Documento.objects.create(
            imovel=self.imovel,
            tipo=self.tipo_matricula,
            numero="M700",
            data=timezone.now().date(),
            cartorio=self.cartorio,
            livro="1",
            folha="1",
        )
        # bulk_create para não disparar o signal de processamento de origens.
        # cartorio_transmissao populado -> exercita a coluna 10 da exportação.
        Lancamento.objects.bulk_create([
            Lancamento(
                documento=self.documento,
                tipo=self.tipo_inicio,
                data=timezone.now().date(),
                cartorio_origem=self.cartorio,
                cartorio_transmissao=self.cartorio_transmissao,
                origem="",
            ),
        ])

    def _request(self, path):
        request = self.factory.get(path)
        request.user = SimpleNamespace(is_authenticated=True)
        return request

    # ------------------------------------------------------------------ Excel

    def _abrir_planilha(self):
        response = cadeia_dominial_views.exportar_cadeia_dominial_excel.__wrapped__(
            self._request("/excel/"), self.tis.id, self.imovel.id
        )
        self.assertEqual(response.status_code, 200)
        return load_workbook(BytesIO(response.content)).active

    def _linha_cabecalho_detalhado(self, ws):
        for linha in ws.iter_rows():
            if linha and linha[0].value == "Nº":
                return linha
        self.fail("Linha de cabeçalho detalhado (começando em 'Nº') não encontrada")

    def test_excel_cabecalhos_das_colunas_4_e_10_sao_cri(self):
        ws = self._abrir_planilha()
        header = self._linha_cabecalho_detalhado(ws)

        # Coluna 4 (cartório da matrícula) e coluna 10 (cartório da transmissão).
        self.assertEqual(header[3].value, "CRI")
        self.assertEqual(header[9].value, "CRI")

        # Nenhuma célula de cabeçalho detalhado menciona "Cartório".
        for cell in header:
            if isinstance(cell.value, str):
                self.assertNotIn("Cartório", cell.value)

    def test_excel_valores_das_colunas_de_cartorio_usam_a_sigla(self):
        ws = self._abrir_planilha()
        header = self._linha_cabecalho_detalhado(ws)
        linha_dados = header[0].row + 1

        self.assertEqual(ws.cell(row=linha_dados, column=4).value, "CRI de Prado")
        self.assertEqual(
            ws.cell(row=linha_dados, column=10).value, "CRI de Teodoro Sampaio"
        )

    def test_excel_nao_deixa_cartorio_por_extenso_fora_do_bloco_do_imovel(self):
        ws = self._abrir_planilha()

        # B7 (info básica do imóvel) é a única exceção: nome por extenso.
        # A7 é o rótulo "Cartório:" desse mesmo bloco.
        self.assertEqual(ws["B7"].value, self.NOME_CRI_IMOVEL)

        for linha in ws.iter_rows():
            for cell in linha:
                if isinstance(cell.value, str) and cell.coordinate not in ("A7", "B7"):
                    self.assertNotIn(
                        "Cartório",
                        cell.value,
                        msg=f"{cell.coordinate} contém 'Cartório': {cell.value!r}",
                    )

    # -------------------------------------------------------------------- PDF

    def _render_pdf(self):
        contexto = CadeiaCompletaService().get_cadeia_completa(
            self.tis.id, self.imovel.id
        )
        return render_to_string("dominial/cadeia_completa_pdf.html", contexto)

    def test_pdf_cabecalhos_usam_cri_e_nao_cartorio(self):
        html = self._render_pdf()
        self.assertIn("<th>CRI</th>", html)
        self.assertNotIn("<th>Cartório</th>", html)

    def test_pdf_valores_tabulares_usam_a_sigla(self):
        html = self._render_pdf()
        self.assertIn("CRI de Prado", html)
        self.assertIn("CRI de Teodoro Sampaio", html)

    def test_pdf_bloco_do_imovel_preserva_nome_por_extenso(self):
        html = self._render_pdf()
        # O bloco "CRI:" das informações do imóvel mantém o nome completo.
        #
        # `assertIn` genérico seria fraco: o mesmo nome por extenso também
        # aparece no cabeçalho narrativo de cada documento
        # (<small>CRI: ...</small>), então a remoção acidental deste bloco não
        # quebraria o teste. Validamos a estrutura exata da linha do bloco
        # `imovel-info` e exigimos que ela ocorra uma única vez.
        self.assertInHTML(
            '<div class="imovel-info-row">'
            '<div class="imovel-info-cell">CRI:</div>'
            f'<div class="imovel-info-cell">{self.NOME_CRI_IMOVEL}</div>'
            "</div>",
            html,
            count=1,
        )
