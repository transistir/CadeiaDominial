from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from django.conf import settings
from django.core.cache import cache
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from dominial.models import (
    Cartorios,
    Documento,
    DocumentoTipo,
    Imovel,
    Lancamento,
    LancamentoTipo,
    OrigemFimCadeia,
    Pessoas,
    TIs,
)
from dominial.services.cadeia_completa_service import CadeiaCompletaService
from dominial.services.hierarquia_arvore_service import HierarquiaArvoreService
from dominial.views import cadeia_dominial_views


class TipoDocumentoFake:
    def __init__(self, tipo, display):
        self.tipo = tipo
        self._display = display

    def get_tipo_display(self):
        return self._display


class ExportacaoCadeiaParidadeTest(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.tis = SimpleNamespace(id=10, nome="TI Teste")
        self.imovel = SimpleNamespace(
            id=20,
            nome="Imóvel Teste",
            matricula="100",
            proprietario=SimpleNamespace(nome="Proprietário Teste"),
            cartorio=SimpleNamespace(nome="Cartório Teste"),
        )
        self.documentos = [
            SimpleNamespace(
                id=101,
                numero="M100",
                tipo=TipoDocumentoFake("matricula", "Matrícula"),
            ),
            SimpleNamespace(
                id=202,
                numero="T90",
                tipo=TipoDocumentoFake("transcricao", "Transcrição"),
            ),
        ]
        self.contexto_completo = {
            "tis": self.tis,
            "imovel": self.imovel,
            "cadeia_completa": [
                {
                    "tipo": "tronco_principal",
                    "titulo": "Tronco principal",
                    "documentos": [
                        {
                            "documento": self.documentos[0],
                            "lancamentos": [],
                            "is_importado": False,
                        },
                        {
                            "documento": self.documentos[1],
                            "lancamentos": [],
                            "is_importado": True,
                        },
                    ],
                }
            ],
            "estatisticas": {
                "total_documentos": 2,
                "total_lancamentos": 0,
                "documentos_importados": 1,
                "total_troncos": 1,
            },
        }

    def _request(self, path, query=None):
        request = self.factory.get(path, data=query or {})
        request.user = SimpleNamespace(is_authenticated=True)
        return request

    def test_botao_pdf_padrao_aponta_para_exportacao_completa(self):
        template = (
            Path(settings.BASE_DIR)
            / "templates"
            / "dominial"
            / "cadeia_dominial_tabela.html"
        ).read_text(encoding="utf-8")

        link_esperado = (
            "{% url 'exportar_cadeia_completa_pdf' "
            "tis_id=tis.id imovel_id=imovel.id %}"
        )
        link_antigo = (
            "{% url 'exportar_cadeia_dominial_pdf' "
            "tis_id=tis.id imovel_id=imovel.id %}"
        )

        self.assertIn(link_esperado, template)
        self.assertNotIn(link_antigo, template)

    def test_rota_antiga_de_pdf_permanece_disponivel(self):
        self.assertEqual(
            reverse(
                "exportar_cadeia_dominial_pdf",
                kwargs={"tis_id": self.tis.id, "imovel_id": self.imovel.id},
            ),
            "/dominial/tis/10/imovel/20/cadeia-tabela/pdf/",
        )

    def test_weasyprint_instalado_gera_pdf_real(self):
        pdf = cadeia_dominial_views.HTML(
            string='<meta charset="utf-8"><h1>Teste de geração PDF</h1>'
        ).write_pdf()

        self.assertGreater(len(pdf), 1000)
        self.assertEqual(pdf[:4], b"%PDF")

    def test_estatisticas_informam_total_de_troncos(self):
        cadeia_completa = [
            {
                "tipo": "tronco_principal",
                "documentos": self.contexto_completo["cadeia_completa"][0][
                    "documentos"
                ],
            },
            {
                "tipo": "tronco_secundario",
                "documentos": [],
            },
        ]

        estatisticas = CadeiaCompletaService()._calcular_estatisticas_completas(
            cadeia_completa
        )

        self.assertEqual(estatisticas["total_troncos"], 2)
        self.assertEqual(estatisticas["total_documentos"], 2)

    @patch("dominial.services.cadeia_completa_service.CadeiaCompletaService")
    @patch.object(cadeia_dominial_views, "get_object_or_404")
    @patch.object(cadeia_dominial_views, "render_to_string")
    @patch.object(cadeia_dominial_views, "HTML")
    @patch.object(cadeia_dominial_views.os.path, "exists", return_value=True)
    def test_pdf_completo_recebe_contexto_e_ordem_do_servico(
        self,
        _exists,
        html_mock,
        render_to_string_mock,
        get_object_mock,
        service_class_mock,
    ):
        get_object_mock.side_effect = [self.tis, self.imovel]
        service = service_class_mock.return_value
        service.get_cadeia_completa.return_value = self.contexto_completo
        render_to_string_mock.return_value = "<html></html>"
        html_mock.return_value.write_pdf.return_value = b"%PDF-teste"

        response = cadeia_dominial_views.exportar_cadeia_completa_pdf.__wrapped__(
            self._request("/pdf/"), self.tis.id, self.imovel.id
        )

        service.get_cadeia_completa.assert_called_once_with(self.tis.id, self.imovel.id)
        render_to_string_mock.assert_called_once_with(
            "dominial/cadeia_completa_pdf.html", self.contexto_completo
        )
        contexto_pdf = render_to_string_mock.call_args.args[1]
        ids_pdf = [
            item["documento"].id
            for tronco in contexto_pdf["cadeia_completa"]
            for item in tronco["documentos"]
        ]
        self.assertEqual(ids_pdf, [101, 202])
        self.assertEqual(response["Content-Type"], "application/pdf")

    @patch("dominial.services.cadeia_completa_service.CadeiaCompletaService")
    @patch.object(cadeia_dominial_views, "get_object_or_404")
    def test_excel_usa_mesmo_servico_e_preserva_ordem_dos_documentos(
        self, get_object_mock, service_class_mock
    ):
        get_object_mock.side_effect = [self.tis, self.imovel]
        service = service_class_mock.return_value
        service.get_cadeia_completa.return_value = self.contexto_completo

        response = cadeia_dominial_views.exportar_cadeia_dominial_excel.__wrapped__(
            self._request("/excel/"), self.tis.id, self.imovel.id
        )

        service.get_cadeia_completa.assert_called_once_with(self.tis.id, self.imovel.id)
        workbook = load_workbook(BytesIO(response.content))
        valores_coluna_a = [cell.value for cell in workbook.active["A"]]
        titulos_esperados = ["Matrícula: M100", "📥 Transcrição: T90"]
        titulos_documentos = [valor for valor in valores_coluna_a if valor in titulos_esperados]
        self.assertEqual(titulos_documentos, titulos_esperados)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @patch("dominial.services.cadeia_completa_service.CadeiaCompletaService")
    @patch.object(cadeia_dominial_views, "get_object_or_404")
    @patch.object(cadeia_dominial_views, "render_to_string", return_value="<html></html>")
    @patch.object(cadeia_dominial_views, "HTML")
    @patch.object(cadeia_dominial_views.os.path, "exists", return_value=True)
    def test_pdf_com_sequencia_preserva_fluxo_personalizado(
        self,
        _exists,
        html_mock,
        _render_to_string_mock,
        get_object_mock,
        service_class_mock,
    ):
        get_object_mock.side_effect = [self.tis, self.imovel]
        service = service_class_mock.return_value
        service.get_cadeia_completa_com_sequencia_personalizada.return_value = (
            self.contexto_completo
        )
        html_mock.return_value.write_pdf.return_value = b"%PDF-teste"

        cadeia_dominial_views.exportar_cadeia_completa_pdf.__wrapped__(
            self._request("/pdf/", {"sequencia": "202,101"}),
            self.tis.id,
            self.imovel.id,
        )

        service.get_cadeia_completa_com_sequencia_personalizada.assert_called_once_with(
            self.tis.id, self.imovel.id, "202,101"
        )
        service.get_cadeia_completa.assert_not_called()


class ExportacaoCadeiaComFimCadeiaTest(TestCase):
    """
    Regressão da issue #146 (banco de dados real).

    `ExportacaoCadeiaParidadeTest`, acima, usa apenas `SimpleNamespace` e
    nunca toca o banco — por isso não pegou este bug. Aqui montamos uma
    cadeia real (matrícula -> transcrição de origem) com uma origem de fim
    de cadeia (issue #85), que faz `HierarquiaArvoreService` injetar em
    `arvore['documentos']` um nó sintético (dict puro, sem Documento por
    trás) cujo id é a string `fim_cadeia_{doc_id}_{lanc_id}_{origem_id}`.

    `CadeiaCompletaService._obter_tronco_principal_completo` fazia
    `Documento.objects.get(id=doc_node['id'])` para todo nó da árvore,
    inclusive o sintético, e explodia com:
        ValueError: Field 'id' expected a number but got 'fim_cadeia_...'
    derrubando com HTTP 500 tanto a exportação em Excel quanto o PDF padrão.
    """

    def setUp(self):
        # LocMemCache é compartilhado entre métodos de teste no mesmo
        # processo (só o banco é revertido a cada teste); sem isso, o cache
        # de tronco principal por imovel_id poderia vazar entre os testes
        # desta classe, já que os IDs são reaproveitados a cada rollback.
        cache.clear()
        self.factory = RequestFactory()

        self.tis = TIs.objects.create(
            nome="TI Teste 146", codigo="TI146", etnia="Teste"
        )
        self.cartorio = Cartorios.objects.create(
            nome="Cartório Teste 146", cns="146146", cidade="Cidade", estado="TS"
        )
        self.proprietario = Pessoas.objects.create(
            nome="Proprietário 146", cpf="11122233344"
        )
        self.imovel = Imovel.objects.create(
            terra_indigena_id=self.tis,
            nome="Imóvel Teste 146",
            proprietario=self.proprietario,
            matricula="M500",
            cartorio=self.cartorio,
        )

        self.tipo_matricula = DocumentoTipo.objects.create(tipo='matricula')
        self.tipo_transcricao = DocumentoTipo.objects.create(tipo='transcricao')
        self.tipo_inicio = LancamentoTipo.objects.create(tipo='inicio_matricula')

        # Documento 1 (nível 0): matrícula que é a identidade registral do
        # imóvel — é o ponto de partida de HierarquiaArvoreService.
        self.documento_matricula = Documento.objects.create(
            imovel=self.imovel,
            tipo=self.tipo_matricula,
            numero="M500",
            data=timezone.now().date(),
            cartorio=self.cartorio,
            livro="1",
            folha="1",
        )

        # Documento 2 (nível 1): transcrição de origem que será referenciada
        # pelo início de matrícula abaixo via identidade (tipo + número +
        # cartório) — cadeia real, encadeada de fato, com mais de um documento.
        self.documento_transcricao = Documento.objects.create(
            imovel=self.imovel,
            tipo=self.tipo_transcricao,
            numero="T90",
            data=timezone.now().date(),
            cartorio=self.cartorio,
            livro="2",
            folha="5",
        )

        # Início de matrícula do documento 1 aponta para "T90" (fallback
        # textual lido por LancamentoOrigemLeituraService, sem
        # LancamentoOrigem estruturada). Usa bulk_create para não disparar o
        # signal post_save de Lancamento (processar_origens_automaticas_signal,
        # dominial/signals.py) — ele criaria automaticamente um segundo
        # Documento "T90" e colidiria com o que acabamos de criar acima
        # (mesmo padrão de test_divida_cartorio_arbitrario_arvore.py).
        Lancamento.objects.bulk_create([
            Lancamento(
                documento=self.documento_matricula,
                tipo=self.tipo_inicio,
                data=timezone.now().date(),
                cartorio_origem=self.cartorio,
                origem="T90",
            ),
        ])
        # Lançamento da transcrição sem origem própria: é nele que a origem
        # de fim de cadeia (issue #85) é registrada, encerrando a cadeia.
        self.lancamento_transcricao = Lancamento.objects.create(
            documento=self.documento_transcricao,
            tipo=self.tipo_inicio,
            data=timezone.now().date(),
            valor_transacao=1000.00,
            origem="",
        )
        self.origem_fim_cadeia = OrigemFimCadeia.objects.create(
            lancamento=self.lancamento_transcricao,
            indice_origem=0,
            fim_cadeia=True,
            tipo_fim_cadeia='destacamento_publico',
            classificacao_fim_cadeia='origem_lidima',
        )

    def _request(self, path):
        request = self.factory.get(path)
        request.user = SimpleNamespace(is_authenticated=True)
        return request

    def test_fixture_gera_no_sintetico_fim_cadeia_na_arvore(self):
        """
        Prova que a fixture realmente aciona a issue #85: sem isto, os
        testes de regressão abaixo seriam vazios e passariam mesmo com o
        bug presente (o problema da classe `ExportacaoCadeiaParidadeTest`).
        """
        arvore = HierarquiaArvoreService.construir_arvore_cadeia_dominial(self.imovel)

        nos_fim_cadeia = [d for d in arvore['documentos'] if d.get('is_fim_cadeia')]
        self.assertEqual(len(nos_fim_cadeia), 1)

        no_fc = nos_fim_cadeia[0]
        self.assertIsInstance(no_fc['id'], str)
        self.assertEqual(
            no_fc['id'],
            f"fim_cadeia_{self.documento_transcricao.id}_"
            f"{self.lancamento_transcricao.id}_{self.origem_fim_cadeia.id}",
        )

        # E a árvore contém os dois documentos reais da cadeia, além do nó sintético.
        ids_reais = {
            d['id'] for d in arvore['documentos'] if not d.get('is_fim_cadeia')
        }
        self.assertEqual(
            ids_reais, {self.documento_matricula.id, self.documento_transcricao.id}
        )

    def test_get_cadeia_completa_nao_lanca_valueerror(self):
        """
        Regressão direta da issue #146: o nó sintético de fim de cadeia não
        deve derrubar `CadeiaCompletaService.get_cadeia_completa` com
        `ValueError: Field 'id' expected a number but got 'fim_cadeia_...'`.
        """
        resultado = CadeiaCompletaService().get_cadeia_completa(
            self.tis.id, self.imovel.id
        )

        self.assertIn('cadeia_completa', resultado)
        self.assertEqual(resultado['estatisticas']['total_documentos'], 2)

    def test_cadeia_completa_contem_apenas_documentos_reais(self):
        """
        A cadeia completa deve conter exatamente os documentos reais criados
        (M500 e T90), todos instâncias de `Documento` com id inteiro, e
        nenhuma entrada sintética de fim de cadeia.
        """
        resultado = CadeiaCompletaService().get_cadeia_completa(
            self.tis.id, self.imovel.id
        )

        documentos = [
            item['documento']
            for tronco in resultado['cadeia_completa']
            for item in tronco['documentos']
        ]

        self.assertEqual(len(documentos), 2)
        for documento in documentos:
            self.assertIsInstance(documento, Documento)
            self.assertIsInstance(documento.id, int)

        self.assertEqual(
            {documento.numero for documento in documentos}, {"M500", "T90"}
        )
        self.assertEqual(
            {documento.id for documento in documentos},
            {self.documento_matricula.id, self.documento_transcricao.id},
        )

    def test_excel_export_retorna_200_com_no_fim_cadeia(self):
        """
        Regressão da issue #146: o botão de exportação Excel não deve mais
        devolver HTTP 500 (`Erro ao gerar Excel: ...`, text/plain) quando a
        árvore contém um nó sintético de fim de cadeia.
        """
        response = cadeia_dominial_views.exportar_cadeia_dominial_excel.__wrapped__(
            self._request("/excel/"), self.tis.id, self.imovel.id
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_excel_export_abre_no_openpyxl_com_documentos_reais(self):
        """
        O arquivo XLSX gerado deve abrir corretamente no openpyxl e conter
        as linhas de título dos dois documentos reais da cadeia.
        """
        response = cadeia_dominial_views.exportar_cadeia_dominial_excel.__wrapped__(
            self._request("/excel/"), self.tis.id, self.imovel.id
        )
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.content))
        valores_coluna_a = [cell.value for cell in workbook.active["A"]]

        titulos_esperados = ["Matrícula: M500", "Transcrição: T90"]
        titulos_documentos = [
            valor for valor in valores_coluna_a if valor in titulos_esperados
        ]
        self.assertEqual(titulos_documentos, titulos_esperados)

        # Nenhuma linha de documento deve corresponder ao nó sintético de
        # fim de cadeia (ele não é um Documento e não deve virar uma linha).
        self.assertFalse(
            any(
                isinstance(valor, str) and "Fim de Cadeia" in valor
                for valor in valores_coluna_a
            )
        )
